"""Recover Sasidharan 2023 FVOC XLSX through the source-issued signed CDN route.

The URL below was resolved from the article's Supplementary data Table S1 link.
This script preserves the workbook and exports all sheets without interpretation.
"""
from __future__ import annotations
import argparse, csv, hashlib, json, re, subprocess
from datetime import datetime, timezone
from pathlib import Path

SIGNED_URL = "https://oup.silverchair-cdn.com/oup/backfile/Content_public/Journal/aob/132/1/10.1093_aob_mcad064/2/mcad064_suppl_supplementary_data.xlsx?Expires=2147483647&Key-Pair-Id=APKAIE5G5CRDK6RD3PGA&Signature=K8CpA8Ps4LKtiGoz1KhgRQy~g3RSfx1xnP5DJVXK-zZRGnnVs~q2d~thDWMnKqLLRHruhD3gEj9Wu0-RPwIMOiuPW5P6VDFoHdZ1OdY9v4vS9CoVXgiZ27cxcYNIbtZ6S1K0SPqFgXb9T~dyNLmRlDK~JhGRLPxYofAs1Gv5DQVs8r9eVQ7VCarkkExfW54J960KC1UvCwbDH7Y7rlgLZPF~HJ7ewkKAJbVZZ0g6Nm3Q7RlygM5x6ph3f3KXsR12muKbYGOnx4Nz~ex~2wxfksXIXSE6dprhKEk4auiSg9vBHEwr7DMPJbyZfVFVv5~BX72alJ0N846lPuajtockuw__"
ARTICLE_URL = "https://academic.oup.com/aob/article/132/1/1/7176361"


def download() -> bytes:
    cmd = ["curl","--fail","--location","--silent","--show-error","--compressed","--retry","3",
           "--connect-timeout","20","--max-time","120","--referer",ARTICLE_URL,
           "--user-agent","Mozilla/5.0 bita-source-recovery/2.0",SIGNED_URL]
    run = subprocess.run(cmd, capture_output=True)
    if run.returncode:
        raise RuntimeError(run.stderr.decode("utf-8", errors="replace"))
    if run.stdout[:2] != b"PK":
        raise RuntimeError(f"signed route did not return XLSX: {len(run.stdout)} bytes")
    return run.stdout


def safe(name: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", name).strip("_") or "sheet"


def main(argv=None) -> int:
    from openpyxl import load_workbook
    parser = argparse.ArgumentParser()
    parser.add_argument("output_dir")
    args = parser.parse_args(argv)
    out = Path(args.output_dir); out.mkdir(parents=True, exist_ok=True)
    payload = download()
    xlsx = out / "sasidharan2023_supplement.xlsx"; xlsx.write_bytes(payload)
    wb = load_workbook(xlsx, read_only=True, data_only=False)
    sheets = out / "sheets"; sheets.mkdir(exist_ok=True)
    inventory = []
    for i, ws in enumerate(wb.worksheets, 1):
        target = sheets / f"{i:02d}_{safe(ws.title)}.csv"
        nr = nc = 0
        with target.open("w", encoding="utf-8", newline="") as h:
            w = csv.writer(h)
            for row in ws.iter_rows(values_only=True):
                vals = ["" if v is None else v for v in row]; w.writerow(vals)
                c = sum(v not in ("", None) for v in vals)
                if c: nr += 1; nc += c
        inventory.append({"sheet_index":i,"sheet_name":ws.title,"max_row":ws.max_row,"max_column":ws.max_column,
                          "nonempty_rows":nr,"nonempty_cells":nc,"csv_path":str(target.relative_to(out))})
    with (out/"workbook_inventory.csv").open("w", encoding="utf-8", newline="") as h:
        w = csv.DictWriter(h, fieldnames=list(inventory[0])); w.writeheader(); w.writerows(inventory)
    receipt = {"article_doi":"10.1093/aob/mcad064","retrieved_at_utc":datetime.now(timezone.utc).isoformat(),
               "source_route":"article-resolved signed OUP CDN supplementary-data link","sha256":hashlib.sha256(payload).hexdigest(),
               "bytes":len(payload),"sheet_count":len(inventory),"inventory":inventory,
               "interpretation_boundary":"Source rows only; FVOC x insect rows remain dependent until reconstructed by publication cluster."}
    (out/"source_receipt.json").write_text(json.dumps(receipt, indent=2)+"\n", encoding="utf-8")
    print(json.dumps(receipt, indent=2))
    return 0

if __name__ == "__main__": raise SystemExit(main())
