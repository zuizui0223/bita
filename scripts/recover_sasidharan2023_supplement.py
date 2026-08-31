"""Recover and inventory Sasidharan et al. 2023 FVOC supplementary workbook.

This is source recovery only. It downloads the public XLSX for
DOI 10.1093/aob/mcad064, exports every worksheet losslessly to CSV, and
records provenance. No FVOC/insect row is promoted to an independent study.

Usage:
    python scripts/recover_sasidharan2023_supplement.py OUTPUT_DIR
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import html
import io
import json
import re
import subprocess
import tarfile
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from pathlib import Path

ARTICLE_URLS = (
    "https://academic.oup.com/aob/article/132/1/1/7176361",
    "https://pmc.ncbi.nlm.nih.gov/articles/PMC10550281/",
)
PMC_ATTACHMENT_URLS = (
    "https://pmc.ncbi.nlm.nih.gov/articles/instance/10550281/bin/mcad064_suppl_supplementary_data.xlsx",
    "https://pmc.ncbi.nlm.nih.gov/articles/PMC10550281/bin/mcad064_suppl_supplementary_data.xlsx",
)
PMC_OA_API = "https://www.ncbi.nlm.nih.gov/pmc/utils/oa/oa.fcgi?id=PMC10550281"
FALLBACK_XLSX = (
    "https://oup.silverchair-cdn.com/oup/backfile/Content_public/Journal/aob/132/1/"
    "10.1093_aob_mcad064/2/mcad064_suppl_supplementary_data.xlsx"
)
USER_AGENT = "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 bita-source-recovery/1.0"


def _request(url: str, referer: str | None = None) -> bytes:
    command = ["curl", "--fail", "--location", "--silent", "--show-error", "--compressed",
               "--retry", "3", "--connect-timeout", "20", "--max-time", "120",
               "--user-agent", USER_AGENT,
               "--header", "Accept: text/html,application/xml,application/gzip,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*"]
    if referer:
        command.extend(["--referer", referer])
    command.append(url)
    completed = subprocess.run(command, check=False, capture_output=True)
    if completed.returncode == 0:
        return completed.stdout
    headers = {"User-Agent": USER_AGENT}
    if referer:
        headers["Referer"] = referer
    req = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(req, timeout=60) as response:
        return response.read()


def _candidate_links(page_url: str, body: bytes) -> list[str]:
    text = html.unescape(body.decode("utf-8", errors="replace"))
    raw = re.findall(r"(?:href|src)=[\"']([^\"']+)[\"']", text, flags=re.I)
    out: list[str] = []
    for href in raw:
        absolute = urllib.parse.urljoin(page_url, href.replace("\\/", "/"))
        low = absolute.lower()
        if ".xlsx" in low and ("mcad064" in low or "supplement" in low):
            out.append(absolute)
    return list(dict.fromkeys(out))


def _try_url(url: str, attempts: list[dict[str, object]], referer: str | None = None):
    try:
        payload = _request(url, referer)
    except Exception as exc:  # network dependent
        attempts.append({"url": url, "status": "failed", "error": repr(exc)})
        return None
    if payload[:2] == b"PK":
        attempts.append({"url": url, "status": "xlsx_recovered", "bytes": len(payload)})
        return url, payload
    attempts.append({"url": url, "status": "not_xlsx", "bytes": len(payload)})
    for nested in _candidate_links(url, payload):
        try:
            data = _request(nested, url)
        except Exception as exc:
            attempts.append({"url": nested, "status": "failed", "error": repr(exc)})
            continue
        if data[:2] == b"PK":
            attempts.append({"url": nested, "status": "xlsx_recovered", "bytes": len(data)})
            return nested, data
    return None


def _recover_from_oa_package(attempts: list[dict[str, object]]):
    try:
        xml = _request(PMC_OA_API)
        root = ET.fromstring(xml)
    except Exception as exc:
        attempts.append({"url": PMC_OA_API, "status": "oa_api_failed", "error": repr(exc)})
        return None
    for element in root.iter():
        href = element.attrib.get("href", "")
        if not href or not href.endswith((".tar.gz", ".tgz")):
            continue
        package_url = href.replace("ftp://", "https://")
        try:
            package = _request(package_url, PMC_OA_API)
            with tarfile.open(fileobj=io.BytesIO(package), mode="r:*") as archive:
                for member in archive.getmembers():
                    name = member.name.lower()
                    if member.isfile() and name.endswith(".xlsx") and "mcad064" in name:
                        handle = archive.extractfile(member)
                        if handle:
                            payload = handle.read()
                            if payload[:2] == b"PK":
                                source = f"{package_url}#{member.name}"
                                attempts.append({"url": source, "status": "xlsx_recovered_from_oa_package", "bytes": len(payload)})
                                return source, payload
        except Exception as exc:
            attempts.append({"url": package_url, "status": "oa_package_failed", "error": repr(exc)})
    return None


def discover_xlsx():
    attempts: list[dict[str, object]] = []
    for page in ARTICLE_URLS:
        try:
            body = _request(page)
        except Exception as exc:
            attempts.append({"url": page, "status": "page_failed", "error": repr(exc)})
            continue
        links = _candidate_links(page, body)
        attempts.append({"url": page, "status": "page_retrieved", "candidate_links": links})
        for link in links:
            result = _try_url(link, attempts, page)
            if result:
                return (*result, attempts)
    for url in PMC_ATTACHMENT_URLS:
        result = _try_url(url, attempts, ARTICLE_URLS[1])
        if result:
            return (*result, attempts)
    result = _recover_from_oa_package(attempts)
    if result:
        return (*result, attempts)
    result = _try_url(FALLBACK_XLSX, attempts, ARTICLE_URLS[0])
    if result:
        return (*result, attempts)
    raise RuntimeError(json.dumps({"message": "Sasidharan XLSX not recovered", "attempts": attempts}, indent=2))


def _safe(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("_") or "sheet"


def export_workbook(path: Path, output_dir: Path):
    from openpyxl import load_workbook
    workbook = load_workbook(path, read_only=True, data_only=False)
    sheet_dir = output_dir / "sheets"
    sheet_dir.mkdir(exist_ok=True)
    inventory = []
    for index, ws in enumerate(workbook.worksheets, start=1):
        csv_path = sheet_dir / f"{index:02d}_{_safe(ws.title)}.csv"
        nonempty_rows = nonempty_cells = 0
        with csv_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            for row in ws.iter_rows(values_only=True):
                values = ["" if v is None else v for v in row]
                writer.writerow(values)
                count = sum(v not in ("", None) for v in values)
                if count:
                    nonempty_rows += 1
                    nonempty_cells += count
        inventory.append({"sheet_index": index, "sheet_name": ws.title, "max_row": ws.max_row,
                          "max_column": ws.max_column, "nonempty_rows": nonempty_rows,
                          "nonempty_cells": nonempty_cells, "csv_path": str(csv_path.relative_to(output_dir))})
    return inventory


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("output_dir")
    args = parser.parse_args(argv)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    source_url, payload, attempts = discover_xlsx()
    xlsx_path = output_dir / "sasidharan2023_supplement.xlsx"
    xlsx_path.write_bytes(payload)
    inventory = export_workbook(xlsx_path, output_dir)
    with (output_dir / "workbook_inventory.csv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(inventory[0]))
        writer.writeheader(); writer.writerows(inventory)
    receipt = {
        "article_doi": "10.1093/aob/mcad064",
        "article_title": "Floral volatiles evoke partially similar responses in both florivores and pollinators and are correlated with non-volatile reward chemicals",
        "retrieved_at_utc": datetime.now(timezone.utc).isoformat(),
        "source_url": source_url,
        "sha256": hashlib.sha256(payload).hexdigest(),
        "bytes": len(payload),
        "sheet_count": len(inventory),
        "inventory": inventory,
        "attempts": attempts,
        "interpretation_boundary": "Workbook recovery preserves source rows only. FVOC x insect rows are dependent within publications and are not independent studies until publication-cluster reconstruction is completed."
    }
    (output_dir / "source_receipt.json").write_text(json.dumps(receipt, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(receipt, indent=2))
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
