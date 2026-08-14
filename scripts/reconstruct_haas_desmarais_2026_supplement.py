from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import tarfile
import urllib.parse
import urllib.request
import zipfile
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "empirical" / "mechanism_pattern_synthesis" / "HAAS_DESMARAIS_2026_SUPPLEMENT_RECEIPT_V1.json"
PMCID = "PMC13095882"
ARTICLE_URL = f"https://pmc.ncbi.nlm.nih.gov/articles/{PMCID}/"
OUP_URL = "https://academic.oup.com/aob/article/137/4/879/8287317"
OA_API = f"https://www.ncbi.nlm.nih.gov/pmc/utils/oa/oa.fcgi?id={PMCID}"
TARGET_FRAGMENT = "mcaf258"
# Exact supplementary-data target resolved from the publisher's Supplementary Data link.
DIRECT_CDN_URL = "https://oup.silverchair-cdn.com/oup/backfile/Content_public/Journal/aob/137/4/10.1093_aob_mcaf258/1/mcaf258_supplementary_data.zip?Expires=2147483647&Key-Pair-Id=APKAIE5G5CRDK6RD3PGA&Signature=TxNgkrKxAeHky~DlvAzhR2CcD0BFccVWQyojXtSo9eDwJ0TgTrlE5kEIG7I~4Jo9iLe56Qx8vKn6~lP9PePQeu9GFaJ-IdCaOHObtEfi9wkFZk5x1AEsk~7mbhch9JhmVwgnqTobvtQKkmkVrw35sIz5x3UmHwBemBNgdijh2~usHJmgyuTaTtz6egh3uXbc~~G3NAStYFsD9PIkttLQpp8DL~IwPfyxMwKsx-IrLZKf9ExYzkOyuY0N~PfIxoG4EIgmxjJxc1JImgCwuhyRumVjqhCtvCRynutQpDyYYhPkCtr7bIqQL1L3Io9hwcJ1ljtJcn13pmgqJdaFXIfvsQ__"
UA = "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/126 Safari/537.36 bita-pattern-expansion/1.0"


def fetch(url: str) -> bytes:
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": UA,
            "Accept": "*/*",
            "Referer": OUP_URL,
        },
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        return resp.read()


def inspect_xlsx(blob: bytes) -> dict[str, object]:
    import openpyxl  # type: ignore

    wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        first_rows = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 4), values_only=True):
            first_rows.append([None if v is None else str(v)[:200] for v in row[:25]])
        sheets.append({
            "title": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "first_rows": first_rows,
        })
    return {"sheets": sheets}


def inspect_csv(blob: bytes) -> dict[str, object]:
    text = blob.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = []
    for i, row in enumerate(reader):
        if i < 4:
            rows.append(row[:25])
        else:
            break
    line_count = text.count("\n") + (1 if text and not text.endswith("\n") else 0)
    return {"line_count": line_count, "first_rows": rows}


def inspect_blob(name: str, blob: bytes) -> dict[str, object]:
    suffix = Path(name).suffix.lower()
    item: dict[str, object] = {
        "name": name,
        "size": len(blob),
        "sha256": hashlib.sha256(blob).hexdigest(),
    }
    if suffix == ".xlsx":
        item["xlsx"] = inspect_xlsx(blob)
    elif suffix == ".csv":
        item["csv"] = inspect_csv(blob)
    elif suffix in {".r", ".txt", ".md"}:
        item["text_head"] = blob.decode("utf-8", errors="replace")[:4000]
    elif suffix == ".zip" and blob[:4] == b"PK\x03\x04":
        nested = []
        with zipfile.ZipFile(io.BytesIO(blob)) as zf:
            for info in zf.infolist():
                if not info.is_dir():
                    nested.append(inspect_blob(f"{name}::{info.filename}", zf.read(info.filename)))
        item["nested_zip_entries"] = nested
    return item


def inspect_zip(blob: bytes) -> list[dict[str, object]]:
    entries: list[dict[str, object]] = []
    with zipfile.ZipFile(io.BytesIO(blob)) as zf:
        for info in zf.infolist():
            if not info.is_dir():
                entries.append(inspect_blob(info.filename, zf.read(info.filename)))
    return entries


def retrieve_direct_cdn() -> tuple[bytes, str, str, list[dict[str, str]], list[dict[str, object]]]:
    attempts: list[dict[str, str]] = []
    try:
        blob = fetch(DIRECT_CDN_URL)
        if blob[:4] != b"PK\x03\x04":
            attempts.append({"url": DIRECT_CDN_URL, "state": f"not_zip_prefix:{blob[:40]!r}"})
            raise RuntimeError("publisher CDN returned non-ZIP content")
        entries = inspect_zip(blob)
        attempts.append({"url": DIRECT_CDN_URL, "state": "publisher_cdn_zip_success"})
        return blob, DIRECT_CDN_URL, "publisher_cdn_zip", attempts, entries
    except Exception as exc:
        attempts.append({"url": DIRECT_CDN_URL, "state": f"publisher_cdn_error:{type(exc).__name__}:{exc}"})
        raise RuntimeError(json.dumps(attempts, indent=2))


def oa_package_candidates() -> tuple[list[str], list[dict[str, str]]]:
    attempts: list[dict[str, str]] = []
    urls: list[str] = []
    try:
        xml = fetch(OA_API).decode("utf-8", errors="replace")
        attempts.append({"url": OA_API, "state": "oa_api_success"})
        for href in re.findall(r'href=["\']([^"\']+)["\']', xml, re.I):
            if ".tar.gz" in href:
                # Keep both the publisher-provided FTP URL and an HTTPS variant.
                urls.append(href)
                if href.startswith("ftp://ftp.ncbi.nlm.nih.gov/"):
                    urls.append("https://ftp.ncbi.nlm.nih.gov/" + href.split("ftp://ftp.ncbi.nlm.nih.gov/", 1)[1])
    except Exception as exc:
        attempts.append({"url": OA_API, "state": f"oa_api_error:{type(exc).__name__}:{exc}"})
    return urls, attempts


def retrieve_from_oa_package() -> tuple[bytes, str, str, list[dict[str, str]], list[dict[str, object]]]:
    candidates, attempts = oa_package_candidates()
    for url in candidates:
        try:
            package = fetch(url)
            if package[:2] != b"\x1f\x8b":
                attempts.append({"url": url, "state": f"not_gzip_prefix:{package[:20]!r}"})
                continue
            entries: list[dict[str, object]] = []
            with tarfile.open(fileobj=io.BytesIO(package), mode="r:gz") as tf:
                for member in tf.getmembers():
                    if not member.isfile():
                        continue
                    fh = tf.extractfile(member)
                    if fh is None:
                        continue
                    blob = fh.read()
                    lower = member.name.lower()
                    if any(lower.endswith(ext) for ext in (".zip", ".xlsx", ".xls", ".csv", ".r", ".txt", ".pdf")):
                        entries.append(inspect_blob(member.name, blob))
            if not entries:
                attempts.append({"url": url, "state": "oa_package_opened_but_no_data_like_entries"})
                continue
            attempts.append({"url": url, "state": "oa_package_success"})
            return package, url, "pmc_oa_tar_gz", attempts, entries
        except Exception as exc:
            attempts.append({"url": url, "state": f"oa_package_error:{type(exc).__name__}:{exc}"})
    raise RuntimeError("PMC OA package retrieval did not yield inspectable entries: " + json.dumps(attempts, indent=2))


def main() -> None:
    all_attempts: list[dict[str, str]] = []
    try:
        package, source_url, package_type, attempts, entries = retrieve_direct_cdn()
        all_attempts.extend(attempts)
    except Exception as direct_exc:
        all_attempts.append({"url": DIRECT_CDN_URL, "state": f"direct_route_failed:{type(direct_exc).__name__}:{direct_exc}"})
        package, source_url, package_type, attempts, entries = retrieve_from_oa_package()
        all_attempts.extend(attempts)

    package_sha = hashlib.sha256(package).hexdigest()
    names = [str(e["name"]) for e in entries]
    if not entries:
        raise RuntimeError("Retrieved package had zero inspectable entries")

    payload = {
        "status": "SUPPLEMENT_PACKAGE_RETRIEVED_AND_INSPECTED",
        "article_doi": "10.1093/aob/mcaf258",
        "pmcid": PMCID,
        "retrieval_url": source_url,
        "package_type": package_type,
        "package_size": len(package),
        "package_sha256": package_sha,
        "attempts": all_attempts,
        "entries": entries,
        "boundary": [
            "this receipt verifies deposited supplementary package structure, not the published meta-analysis estimates by itself",
            "herbivory treatment is not the focal floral defence trait D",
            "no effect from this module is rho, iota, kappa, or W_AD",
        ],
    }
    OUT.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps({"status": payload["status"], "package_sha256": package_sha, "entries": names}, indent=2))


if __name__ == "__main__":
    main()
