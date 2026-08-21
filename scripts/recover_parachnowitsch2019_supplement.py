"""Recover and inventory the study-level supplement for Parachnowitsch et al. 2019.

This is a source-recovery utility, not a meta-analysis. It downloads the publicly
listed XLSX supplement for doi:10.1093/aob/mcy132, records a provenance receipt,
and exports every worksheet to CSV without interpreting or reclassifying rows.

Usage:
    python scripts/recover_parachnowitsch2019_supplement.py OUTPUT_DIR
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import html
import io
import json
import re
import subprocess
import tarfile
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

ARTICLE_URLS = (
    "https://academic.oup.com/aob/article/123/2/247/5055672",
    "https://pmc.ncbi.nlm.nih.gov/articles/PMC6344224/",
)
PMC_ATTACHMENT_URLS = (
    "https://pmc.ncbi.nlm.nih.gov/articles/instance/6344224/bin/mcy132_suppl_aob-18212-s03.xlsx",
    "https://pmc.ncbi.nlm.nih.gov/articles/PMC6344224/bin/mcy132_suppl_aob-18212-s03.xlsx",
)
PMC_OA_API = "https://www.ncbi.nlm.nih.gov/pmc/utils/oa/oa.fcgi?id=PMC6344224"
FALLBACK_XLSX = (
    "https://oup.silverchair-cdn.com/oup/backfile/Content_public/Journal/aob/123/2/"
    "10.1093_aob_mcy132/1/mcy132_suppl_aob-18212-s03.xlsx"
)
USER_AGENT = "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 bita-source-recovery/1.2"


def _request_with_curl(url: str, *, referer: str | None = None) -> bytes:
    command = [
        "curl",
        "--fail",
        "--location",
        "--silent",
        "--show-error",
        "--compressed",
        "--retry",
        "3",
        "--connect-timeout",
        "20",
        "--max-time",
        "120",
        "--user-agent",
        USER_AGENT,
        "--header",
        (
            "Accept: text/html,application/xhtml+xml,application/xml,application/gzip,"
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*"
        ),
    ]
    if referer:
        command.extend(["--referer", referer])
    command.append(url)
    completed = subprocess.run(command, check=False, capture_output=True)
    if completed.returncode != 0:
        error = completed.stderr.decode("utf-8", errors="replace").strip()
        raise RuntimeError(f"curl exit {completed.returncode}: {error}")
    return completed.stdout


def _request(url: str, *, referer: str | None = None) -> bytes:
    # PMC currently serves a browser proof-of-work page to urllib-style clients.
    # curl is attempted first because it follows the supported attachment route
    # without treating that HTML challenge as the source file.
    try:
        return _request_with_curl(url, referer=referer)
    except (FileNotFoundError, RuntimeError):
        headers = {
            "User-Agent": USER_AGENT,
            "Accept": (
                "text/html,application/xhtml+xml,application/xml,application/gzip,"
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*"
            ),
        }
        if referer:
            headers["Referer"] = referer
        request = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(request, timeout=60) as response:
            return response.read()


def _candidate_links(page_url: str, body: bytes) -> list[str]:
    text = html.unescape(body.decode("utf-8", errors="replace"))
    raw_links: list[str] = []
    raw_links.extend(re.findall(r"(?:href|src)=[\"']([^\"']+)[\"']", text, flags=re.IGNORECASE))
    raw_links.extend(
        re.findall(
            r"https?://[^\s\"'<>]+(?:\.xlsx(?:\?[^\s\"'<>]*)?|mcy132[^\s\"'<>]*s03[^\s\"'<>]*)",
            text,
            flags=re.IGNORECASE,
        )
    )
    for match in re.findall(r"content=[\"']([^\"']+)[\"']", text, flags=re.IGNORECASE):
        if "url=" in match.lower():
            raw_links.append(re.split(r"url=", match, flags=re.IGNORECASE, maxsplit=1)[1].strip())

    candidates: list[str] = []
    for href in raw_links:
        absolute = urllib.parse.urljoin(page_url, href.replace("\\/", "/"))
        lowered = absolute.lower()
        if (
            ".xlsx" in lowered
            or ("mcy132" in lowered and "s03" in lowered)
            or "cdn.ncbi.nlm.nih.gov/pmc/blobs" in lowered
        ):
            candidates.append(absolute)
    return list(dict.fromkeys(candidates))


def _html_preview(payload: bytes, limit: int = 280) -> str:
    text = html.unescape(payload.decode("utf-8", errors="replace"))
    text = re.sub(r"\s+", " ", text).strip()
    return text[:limit]


def _follow_candidate(
    url: str,
    *,
    attempts: list[dict[str, object]],
    referer: str | None = None,
    seen: set[str] | None = None,
    depth: int = 0,
) -> tuple[str, bytes] | None:
    if seen is None:
        seen = set()
    if url in seen or depth > 3:
        return None
    seen.add(url)

    variants = [url]
    if "pmc.ncbi.nlm.nih.gov" in url and "download=1" not in url:
        separator = "&" if "?" in url else "?"
        variants.append(f"{url}{separator}download=1")

    for candidate in variants:
        if candidate in seen and candidate != url:
            continue
        seen.add(candidate)
        try:
            payload = _request(candidate, referer=referer)
        except Exception as error:  # pragma: no cover - network dependent
            attempts.append({"url": candidate, "status": "download_failed", "error": repr(error)})
            continue
        if payload[:2] == b"PK":
            attempts.append({"url": candidate, "status": "xlsx_recovered", "bytes": len(payload)})
            return candidate, payload

        nested = _candidate_links(candidate, payload)
        attempts.append(
            {
                "url": candidate,
                "status": "not_xlsx",
                "bytes": len(payload),
                "nested_candidate_links": nested,
                "body_preview": _html_preview(payload),
            }
        )
        for link in nested:
            recovered = _follow_candidate(
                link,
                attempts=attempts,
                referer=candidate,
                seen=seen,
                depth=depth + 1,
            )
            if recovered:
                return recovered
    return None


def _recover_from_pmc_package(attempts: list[dict[str, object]]) -> tuple[str, bytes] | None:
    try:
        payload = _request(PMC_OA_API)
    except Exception as error:  # pragma: no cover - network dependent
        attempts.append({"url": PMC_OA_API, "status": "oa_api_failed", "error": repr(error)})
        return None

    attempts.append(
        {
            "url": PMC_OA_API,
            "status": "oa_api_retrieved",
            "bytes": len(payload),
            "body_preview": _html_preview(payload),
        }
    )
    try:
        root = ET.fromstring(payload)
    except ET.ParseError as error:
        attempts.append({"url": PMC_OA_API, "status": "oa_api_invalid_xml", "error": repr(error)})
        return None

    package_urls: list[str] = []
    for element in root.iter():
        href = element.attrib.get("href", "")
        if href and (element.attrib.get("format") in {"tgz", "tar.gz"} or href.endswith((".tar.gz", ".tgz"))):
            package_urls.append(href.replace("ftp://", "https://"))

    for package_url in package_urls:
        try:
            package = _request(package_url, referer=PMC_OA_API)
        except Exception as error:  # pragma: no cover - network dependent
            attempts.append({"url": package_url, "status": "oa_package_failed", "error": repr(error)})
            continue
        attempts.append({"url": package_url, "status": "oa_package_retrieved", "bytes": len(package)})
        try:
            with tarfile.open(fileobj=io.BytesIO(package), mode="r:*") as archive:
                members = [
                    member
                    for member in archive.getmembers()
                    if member.isfile()
                    and (
                        member.name.lower().endswith(".xlsx")
                        or ("mcy132" in member.name.lower() and "s03" in member.name.lower())
                    )
                ]
                for member in members:
                    handle = archive.extractfile(member)
                    if handle is None:
                        continue
                    workbook = handle.read()
                    if workbook[:2] == b"PK":
                        source = f"{package_url}#{member.name}"
                        attempts.append({"url": source, "status": "xlsx_recovered_from_oa_package", "bytes": len(workbook)})
                        return source, workbook
        except (tarfile.TarError, OSError) as error:
            attempts.append({"url": package_url, "status": "oa_package_invalid", "error": repr(error)})
    return None


def discover_xlsx() -> tuple[str, bytes, list[dict[str, object]]]:
    attempts: list[dict[str, object]] = []

    for page_url in ARTICLE_URLS:
        try:
            body = _request(page_url)
        except Exception as error:  # pragma: no cover - network dependent
            attempts.append({"url": page_url, "status": "page_failed", "error": repr(error)})
            continue
        links = _candidate_links(page_url, body)
        attempts.append({"url": page_url, "status": "retrieved", "candidate_links": links})
        for link in links:
            recovered = _follow_candidate(link, attempts=attempts, referer=page_url)
            if recovered:
                source_url, payload = recovered
                return source_url, payload, attempts

    for attachment_url in PMC_ATTACHMENT_URLS:
        recovered = _follow_candidate(
            attachment_url,
            attempts=attempts,
            referer="https://pmc.ncbi.nlm.nih.gov/articles/PMC6344224/",
        )
        if recovered:
            source_url, payload = recovered
            return source_url, payload, attempts

    recovered_package = _recover_from_pmc_package(attempts)
    if recovered_package:
        source_url, payload = recovered_package
        return source_url, payload, attempts

    recovered = _follow_candidate(
        FALLBACK_XLSX,
        attempts=attempts,
        referer="https://academic.oup.com/aob/article/123/2/247/5055672",
    )
    if recovered:
        source_url, payload = recovered
        return source_url, payload, attempts

    raise RuntimeError(json.dumps({"message": "XLSX supplement not recovered", "attempts": attempts}, indent=2))


def _safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("_")
    return cleaned or "sheet"


def _cell_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def export_workbook(xlsx_path: Path, output_dir: Path) -> list[dict[str, object]]:
    from openpyxl import load_workbook

    workbook = load_workbook(xlsx_path, read_only=True, data_only=False)
    inventory: list[dict[str, object]] = []
    sheet_dir = output_dir / "sheets"
    sheet_dir.mkdir(parents=True, exist_ok=True)

    for index, worksheet in enumerate(workbook.worksheets, start=1):
        csv_path = sheet_dir / f"{index:02d}_{_safe_filename(worksheet.title)}.csv"
        nonempty_rows = 0
        nonempty_cells = 0
        with csv_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            for row in worksheet.iter_rows(values_only=True):
                values = [_cell_value(value) for value in row]
                writer.writerow(values)
                row_nonempty = sum(value not in ("", None) for value in values)
                if row_nonempty:
                    nonempty_rows += 1
                    nonempty_cells += row_nonempty
        inventory.append(
            {
                "sheet_index": index,
                "sheet_name": worksheet.title,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "nonempty_rows": nonempty_rows,
                "nonempty_cells": nonempty_cells,
                "csv_path": str(csv_path.relative_to(output_dir)),
            }
        )
    return inventory


def _write_inventory(path: Path, inventory: Iterable[dict[str, object]]) -> None:
    rows = list(inventory)
    fields = (
        "sheet_index",
        "sheet_name",
        "max_row",
        "max_column",
        "nonempty_rows",
        "nonempty_cells",
        "csv_path",
    )
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("output_dir")
    args = parser.parse_args(argv)

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    source_url, payload, attempts = discover_xlsx()
    if payload[:2] != b"PK":
        raise RuntimeError("Recovered supplement does not have XLSX/ZIP magic bytes")

    xlsx_path = output_dir / "parachnowitsch2019_supplement_s3.xlsx"
    xlsx_path.write_bytes(payload)
    digest = hashlib.sha256(payload).hexdigest()
    inventory = export_workbook(xlsx_path, output_dir)
    _write_inventory(output_dir / "workbook_inventory.csv", inventory)

    receipt = {
        "article_doi": "10.1093/aob/mcy132",
        "article_title": "Evolutionary ecology of nectar",
        "retrieved_at_utc": datetime.now(timezone.utc).isoformat(),
        "source_url": source_url,
        "sha256": digest,
        "bytes": len(payload),
        "sheet_count": len(inventory),
        "attempts": attempts,
        "interpretation_boundary": (
            "This recovery preserves the published workbook and worksheet rows. "
            "No row is treated as a strict B-to-pollinator effect until the focal trait role, "
            "outcome lane, effect metric, and study independence are re-audited."
        ),
    }
    (output_dir / "source_receipt.json").write_text(
        json.dumps(receipt, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    print(json.dumps(receipt, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
