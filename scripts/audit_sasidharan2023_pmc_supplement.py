"""Audit the PMC OA supplement for Sasidharan et al. (2023).

Primary article DOI: 10.1093/aob/mcad064, PMCID PMC10550281. The article states
that Tables S1-S5 are the datasets used for its meta-analyses. This script resolves
the NCBI PMC Open Access package, recovers the declared supplementary XLSX in
memory, and writes sheet/header/row-count metadata only. Observation-level values
are never written to repository files or workflow artifacts.

PMC changed its article-dataset distribution in 2026. Legacy `oa_package` objects
were moved under `deprecated/oa_package` in April 2026 ahead of their final removal.
The OA API can therefore expose a legacy-looking package URL whose HTTPS object has
moved. This audit tries the API-declared HTTPS route first and the documented 2026
`deprecated` equivalent second, without changing the requested PMCID or package.
"""

from __future__ import annotations

import io
import json
import re
import tarfile
from pathlib import Path
from urllib.request import Request, urlopen
from xml.etree import ElementTree as ET

PMCID = "PMC10550281"
SOURCE_DOI = "10.1093/aob/mcad064"
OA_API = f"https://www.ncbi.nlm.nih.gov/pmc/utils/oa/oa.fcgi?id={PMCID}"
USER_AGENT = "bita-sasidharan2023-pmc-supplement-audit/1.1"
MAX_PACKAGE_BYTES = 30 * 1024 * 1024
TARGET_TOKEN = "mcad064_suppl_Supplementary_Data"
KEY_TOKENS = (
    "reference", "study", "author", "doi", "plant", "genus", "species", "fvoc", "volatile",
    "compound", "pollinator", "florivore", "insect", "order", "detect", "response", "attract",
    "repel", "behaviour", "behavior", "electro", "pollen", "protein", "toxin", "shannon",
)


def _get(url: str, *, accept: str = "*/*", max_bytes: int = MAX_PACKAGE_BYTES) -> bytes:
    req = Request(url, headers={"User-Agent": USER_AGENT, "Accept": accept})
    with urlopen(req, timeout=60) as response:  # nosec B310: fixed NCBI public endpoints
        length = response.headers.get("Content-Length")
        if length and int(length) > max_bytes:
            raise ValueError(f"response exceeds configured byte limit: {length}")
        data = response.read(max_bytes + 1)
    if len(data) > max_bytes:
        raise ValueError("response exceeded configured byte limit")
    return data


def _oa_package_url() -> str:
    root = ET.fromstring(_get(OA_API, accept="application/xml,text/xml,*/*", max_bytes=1024 * 1024))
    for link in root.findall(".//link"):
        if link.attrib.get("format", "").lower() == "tgz" and link.attrib.get("href"):
            return link.attrib["href"]
    raise RuntimeError("PMC OA API did not expose a tgz package link")


def _normalise_ftp(url: str) -> str:
    # urllib in GitHub Actions is more reliable over HTTPS than FTP. NCBI's OA
    # service commonly returns ftp://ftp.ncbi.nlm.nih.gov/... links.
    if url.startswith("ftp://ftp.ncbi.nlm.nih.gov/"):
        return "https://ftp.ncbi.nlm.nih.gov/" + url[len("ftp://ftp.ncbi.nlm.nih.gov/"):]
    return url


def _package_candidates(api_url: str) -> list[str]:
    """Return current and documented-2026 legacy package routes in priority order."""
    https_url = _normalise_ftp(api_url)
    candidates = [https_url]
    legacy_token = "/pub/pmc/oa_package/"
    deprecated_token = "/pub/pmc/deprecated/oa_package/"
    if legacy_token in https_url:
        candidates.append(https_url.replace(legacy_token, deprecated_token, 1))
    elif deprecated_token in https_url:
        candidates.append(https_url.replace(deprecated_token, legacy_token, 1))
    # Stable de-duplication preserves API route first.
    return list(dict.fromkeys(candidates))


def _download_package(api_url: str) -> tuple[bytes, str, list[dict[str, str]]]:
    attempts: list[dict[str, str]] = []
    for url in _package_candidates(api_url):
        try:
            package = _get(url, accept="application/gzip,application/octet-stream,*/*")
            attempts.append({"url": url, "status": "success"})
            return package, url, attempts
        except Exception as error:
            attempts.append({"url": url, "status": f"{type(error).__name__}: {error}"})
    raise RuntimeError(f"all PMC package routes failed: {attempts}")


def _supplement_from_package(package: bytes) -> tuple[str, bytes, list[str]]:
    archive = tarfile.open(fileobj=io.BytesIO(package), mode="r:gz")
    members = [member for member in archive.getmembers() if member.isfile()]
    names = [member.name for member in members]
    candidates = [
        member for member in members
        if TARGET_TOKEN.lower() in Path(member.name).name.lower()
        and Path(member.name).suffix.lower() == ".xlsx"
    ]
    if len(candidates) != 1:
        raise RuntimeError(f"expected one supplementary XLSX candidate, found {[m.name for m in candidates]}")
    handle = archive.extractfile(candidates[0])
    if handle is None:
        raise RuntimeError("supplementary XLSX could not be extracted in memory")
    return candidates[0].name, handle.read(), names


def _clean_header(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def _audit_xlsx(data: bytes) -> dict[str, object]:
    import openpyxl  # type: ignore

    workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheets: list[dict[str, object]] = []
    for ws in workbook.worksheets:
        row_count = 0
        nonempty_row_count = 0
        first_rows: list[list[str]] = []
        for values in ws.iter_rows(values_only=True):
            row_count += 1
            cleaned = [_clean_header(value) for value in values]
            while cleaned and cleaned[-1] == "":
                cleaned.pop()
            if any(cleaned):
                nonempty_row_count += 1
                if len(first_rows) < 10:
                    # Schema audit only: retain textual strings and suppress numeric
                    # observation values to avoid writing source rows into artifacts.
                    first_rows.append([
                        value if value and not _looks_numeric(value) else "<numeric>"
                        for value in cleaned
                    ])
        header_row_index, headers = _detect_header(first_rows)
        sheets.append({
            "sheet_name": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "iterated_row_count": row_count,
            "nonempty_row_count": nonempty_row_count,
            "detected_header_row_within_first_10_nonempty": header_row_index,
            "headers": headers,
            "candidate_columns": [
                header for header in headers
                if any(token in header.lower() for token in KEY_TOKENS)
            ],
            "first_10_nonempty_schema_rows_numeric_suppressed": first_rows,
        })
    return {"sheet_count": len(sheets), "sheets": sheets}


def _looks_numeric(value: str) -> bool:
    try:
        float(value)
        return True
    except ValueError:
        return False


def _detect_header(rows: list[list[str]]) -> tuple[int | None, list[str]]:
    best: tuple[tuple[int, int], int, list[str]] | None = None
    for index, row in enumerate(rows, start=1):
        strings = [value for value in row if value and value != "<numeric>"]
        if not strings:
            continue
        token_hits = sum(any(token in value.lower() for token in KEY_TOKENS) for value in strings)
        score = (token_hits, len(strings))
        if best is None or score > best[0]:
            best = (score, index, strings)
    return (None, []) if best is None else (best[1], best[2])


def run(output_path: str | Path) -> dict[str, object]:
    oa_url = _oa_package_url()
    package, package_url, package_attempts = _download_package(oa_url)
    supplement_name, supplement, member_names = _supplement_from_package(package)
    report = {
        "source_doi": SOURCE_DOI,
        "pmcid": PMCID,
        "oa_api": OA_API,
        "oa_api_declared_package_url": oa_url,
        "oa_package_url_used": package_url,
        "oa_package_attempts": package_attempts,
        "package_member_count": len(member_names),
        "supplement_name": supplement_name,
        "supplement_size_bytes": len(supplement),
        "workbook": _audit_xlsx(supplement),
        "guardrails": [
            "The article identifies Tables S1-S5 as datasets used in the meta-analyses.",
            "No observation-level numeric values are emitted by this schema audit.",
            "No bita re-analysis is fit until the response unit, dependence structure, and theory-facing estimands are predeclared from the primary article and audited headers.",
            "The package fallback follows NCBI's documented 2026 migration of legacy oa_package objects under deprecated/oa_package; it does not substitute a different article or dataset.",
        ],
    }
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(report, indent=2, sort_keys=True), encoding="utf-8")
    return report


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("output")
    args = parser.parse_args()
    result = run(args.output)
    print(json.dumps({
        "supplement_name": result["supplement_name"],
        "sheet_count": result["workbook"]["sheet_count"],
        "package_url_used": result["oa_package_url_used"],
    }, indent=2))